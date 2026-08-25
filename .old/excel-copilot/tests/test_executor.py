"""Integration tests for ActionExecutor."""

import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import pytest

from excel_copilot.analyzer.workbook_analyzer import WorkbookAnalyzer
from excel_copilot.models.protocol import ActionProtocol, ActionIntent, ActionItem, ExecutionStatus
from excel_copilot.executor.executor_main import ActionExecutor


@pytest.fixture
def sample_excel_file(tmp_path):
    file_path = str(tmp_path / "sales_exec.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    ws.append(["Date", "Revenue", "Cost"])
    ws.append(["2026-01-01", 1000, 600])
    ws.append(["2026-01-02", 1500, 800])

    tab = Table(displayName="SalesData", ref="A1:C3")
    ws.add_table(tab)

    wb.save(file_path)
    wb.close()
    return file_path


def test_executor_add_column_and_formula(sample_excel_file):
    model = WorkbookAnalyzer.analyze(sample_excel_file)
    executor = ActionExecutor(backup_dir=os.path.join(os.path.dirname(sample_excel_file), ".backups"))

    protocol = ActionProtocol(
        intent=ActionIntent.MODIFY_WORKBOOK,
        actions=[
            ActionItem(action="add_column", table="SalesData", column="Profit"),
            ActionItem(action="insert_formula", table="SalesData", column="Profit", formula="{Revenue}-{Cost}"),
            ActionItem(action="create_chart", sheet="Sales", chart="Monthly Sales Chart", chart_type="column"),
        ],
    )

    result, updated_model = executor.execute(protocol, sample_excel_file, model=model)

    assert result.status == ExecutionStatus.SUCCESS
    assert result.actions_executed == 3
    assert updated_model is not None

    # Inspect modified openpyxl workbook
    wb = openpyxl.load_workbook(sample_excel_file)
    ws = wb["Sales"]

    assert ws.cell(row=1, column=4).value == "Profit"
    assert ws.cell(row=2, column=4).value in ("=B2-C2", "=[@Revenue]-[@Cost]")
    assert len(ws._charts) == 1
    wb.close()
