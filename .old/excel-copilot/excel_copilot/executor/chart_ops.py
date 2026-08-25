"""Chart operations: create, update, delete openpyxl charts."""

from typing import Optional
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import range_boundaries, get_column_letter


class ChartOps:
    """Executes chart modifications on openpyxl Workbook."""

    @classmethod
    def execute_create_chart(
        cls,
        wb: openpyxl.Workbook,
        sheet_name: str,
        chart_title: str = "Chart",
        chart_type: str = "bar",
        data_range: Optional[str] = None,
        anchor: str = "E2",
    ) -> str:
        """Create and embed a new openpyxl chart."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        c_type = str(chart_type).lower().strip()
        if "line" in c_type:
            chart = LineChart()
        elif "pie" in c_type:
            chart = PieChart()
        else:
            chart = BarChart()
            chart.type = "col" if "col" in c_type else "bar"

        chart.title = chart_title
        chart.style = 10

        # Define data source reference
        if data_range and ":" in data_range:
            min_c, min_r, max_c, max_r = range_boundaries(data_range)
            data = Reference(ws, min_col=min_c, min_row=min_r, max_col=max_c, max_row=max_r)
            chart.add_data(data, titles_from_data=True)
        else:
            # Default reference top populated area
            max_c = ws.max_column or 2
            max_r = ws.max_row or 5
            data = Reference(ws, min_col=1, min_row=1, max_col=max_c, max_row=max_r)
            chart.add_data(data, titles_from_data=True)

        ws.add_chart(chart, anchor)
        return f"Created {c_type.capitalize()} chart '{chart_title}' anchored at {anchor} in '{ws.title}'."

    @classmethod
    def execute_delete_chart(cls, wb: openpyxl.Workbook, sheet_name: str, chart_name: str) -> str:
        """Remove a chart by title/name."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        if not hasattr(ws, "_charts") or not ws._charts:
            return f"No charts found in sheet '{ws.title}'."

        removed_count = 0
        charts_to_keep = []
        for c in ws._charts:
            title_str = str(c.title) if c.title else ""
            if chart_name.lower() in title_str.lower():
                removed_count += 1
            else:
                charts_to_keep.append(c)

        ws._charts = charts_to_keep
        return f"Deleted {removed_count} chart(s) matching '{chart_name}' from '{ws.title}'."
