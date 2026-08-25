"""Formatting & Style operations: autofit columns, styling headers, freeze panes."""

from typing import Optional, Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class StyleOps:
    """Executes formatting and visual styling on openpyxl Workbook."""

    @classmethod
    def execute_autofit_columns(cls, wb: openpyxl.Workbook, sheet_name: str) -> str:
        """Autofit all column widths in a worksheet."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            # Set width with minimum padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        return f"Autofitted columns for sheet '{ws.title}'."

    @classmethod
    def execute_apply_style(
        cls,
        wb: openpyxl.Workbook,
        sheet_name: str,
        target: str = "headers",
        bg_color: Optional[str] = "1F4E78",  # Default dark blue header
        text_color: Optional[str] = "FFFFFF",
        number_format: Optional[str] = None,
        bold: bool = True,
    ) -> str:
        """Apply visual formatting to headers or cell ranges."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        if target == "headers":
            header_row = 1
            font = Font(bold=bold, color=text_color, name="Calibri", size=11)
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            align = Alignment(horizontal="center", vertical="center")

            for c in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=header_row, column=c)
                cell.font = font
                cell.fill = fill
                cell.alignment = align

            return f"Applied header styling to sheet '{ws.title}'."

        elif number_format:
            # Apply number format across data rows
            for r in range(2, (ws.max_row or 1) + 1):
                for c in range(1, (ws.max_column or 1) + 1):
                    ws.cell(row=r, column=c).number_format = number_format

            return f"Applied number format '{number_format}' to '{ws.title}' data cells."

        return f"Applied custom style to '{ws.title}'."

    @classmethod
    def execute_freeze_panes(cls, wb: openpyxl.Workbook, sheet_name: str, cell_ref: str = "A2") -> str:
        """Freeze panes at target cell reference."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws.freeze_panes = cell_ref
        return f"Frozen panes at '{cell_ref}' on sheet '{ws.title}'."
