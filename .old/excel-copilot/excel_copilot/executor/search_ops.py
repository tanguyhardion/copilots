"""Search operations: Query sheets, tables, columns, formulas, and cell text."""

from typing import List, Dict, Any
import openpyxl
from excel_copilot.models.semantic import WorkbookModel


class SearchOps:
    """Executes read-only query operations across openpyxl workbook."""

    @classmethod
    def find_sheet(cls, model: WorkbookModel, query: str) -> List[Dict[str, Any]]:
        """Find worksheets matching query string."""
        results = []
        q = query.lower().strip()
        for ws in model.worksheets:
            if q in ws.name.lower():
                results.append({
                    "sheet_name": ws.name,
                    "tables_count": len(ws.tables),
                    "charts_count": len(ws.charts),
                    "max_row": ws.max_row,
                })
        return results

    @classmethod
    def find_table(cls, model: WorkbookModel, query: str) -> List[Dict[str, Any]]:
        """Find tables matching query string."""
        results = []
        q = query.lower().strip()
        for ws in model.worksheets:
            for tbl in ws.tables:
                if q in tbl.name.lower() or q in ws.name.lower():
                    results.append({
                        "table_name": tbl.name,
                        "worksheet": ws.name,
                        "range": tbl.range,
                        "columns": [c.name for c in tbl.columns],
                        "row_count": tbl.row_count,
                    })
        return results

    @classmethod
    def find_column(cls, model: WorkbookModel, query: str) -> List[Dict[str, Any]]:
        """Find columns matching query string."""
        results = []
        q = query.lower().strip()
        for ws in model.worksheets:
            for tbl in ws.tables:
                for col in tbl.columns:
                    if q in col.name.lower():
                        results.append({
                            "column_name": col.name,
                            "table_name": tbl.name,
                            "worksheet": ws.name,
                            "data_type": col.data_type,
                            "has_formula": col.has_formula,
                        })
        return results

    @classmethod
    def search_text(cls, wb: openpyxl.Workbook, query: str) -> List[Dict[str, Any]]:
        """Search text across all cells in workbook."""
        results = []
        q = str(query).lower().strip()
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value is not None and q in str(cell.value).lower():
                        results.append({
                            "sheet": sheet_name,
                            "coordinate": cell.coordinate,
                            "value": str(cell.value),
                        })
                        if len(results) >= 50:  # Limit results cap
                            return results
        return results
