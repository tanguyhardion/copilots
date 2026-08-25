"""Data operations: column management, row insertion/deletion, value replacements."""

from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries


class DataOps:
    """Executes data level modifications on openpyxl Workbook."""

    @classmethod
    def execute_add_column(
        cls,
        wb: openpyxl.Workbook,
        sheet_name: str,
        column_name: str,
        table_name: Optional[str] = None,
        default_values: Optional[List[Any]] = None,
    ) -> str:
        """Add a new column to a table or worksheet."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        # If native openpyxl table exists
        if table_name and hasattr(ws, "tables") and table_name in ws.tables:
            tbl = ws.tables[table_name]
            tbl_ref = tbl if isinstance(tbl, str) else getattr(tbl, "ref", str(tbl))
            min_c, min_r, max_c, max_r = range_boundaries(tbl_ref)

            new_col_idx = max_c + 1
            # Write header
            ws.cell(row=min_r, column=new_col_idx, value=column_name)

            # Populate default values if provided
            if default_values:
                for idx, val in enumerate(default_values):
                    target_r = min_r + 1 + idx
                    if target_r <= max_r + 1:
                        ws.cell(row=target_r, column=new_col_idx, value=val)

            # Update table reference range
            new_ref = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(new_col_idx)}{max_r}"
            if isinstance(tbl, str):
                ws.tables[table_name] = new_ref
            elif hasattr(tbl, "ref"):
                tbl.ref = new_ref

            return f"Added column '{column_name}' to table '{table_name}' at range {new_ref}."

        else:
            # Fallback to appending column at max_column + 1
            max_c = ws.max_column or 0
            new_col_idx = max_c + 1
            ws.cell(row=1, column=new_col_idx, value=column_name)

            if default_values:
                for idx, val in enumerate(default_values):
                    ws.cell(row=2 + idx, column=new_col_idx, value=val)

            return f"Added column '{column_name}' to sheet '{ws.title}' at column {get_column_letter(new_col_idx)}."

    @classmethod
    def execute_remove_column(
        cls,
        wb: openpyxl.Workbook,
        sheet_name: str,
        column_name: str,
        table_name: Optional[str] = None,
    ) -> str:
        """Remove a column by header name."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        min_r = 1
        max_r = ws.max_row or 1
        min_c = 1
        max_c = ws.max_column or 1

        if table_name and hasattr(ws, "tables") and table_name in ws.tables:
            tbl = ws.tables[table_name]
            min_c, min_r, max_c, max_r = range_boundaries(tbl.ref)

        # Locate column index by header
        col_idx = None
        for c in range(min_c, max_c + 1):
            val = ws.cell(row=min_r, column=c).value
            if val and str(val).strip().lower() == column_name.strip().lower():
                col_idx = c
                break

        if not col_idx:
            raise ValueError(f"Column '{column_name}' not found in sheet '{ws.title}'.")

        ws.delete_cols(col_idx, 1)

        # Update table ref if table existed
        if table_name and hasattr(ws, "tables") and table_name in ws.tables:
            tbl = ws.tables[table_name]
            new_max_c = max(min_c, max_c - 1)
            tbl.ref = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(new_max_c)}{max_r}"

        return f"Removed column '{column_name}' from '{ws.title}'."

    @classmethod
    def execute_rename_column(
        cls,
        wb: openpyxl.Workbook,
        sheet_name: str,
        old_column_name: str,
        new_column_name: str,
    ) -> str:
        """Rename an existing column header."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        renamed = False
        for c in range(1, (ws.max_column or 1) + 1):
            cell = ws.cell(row=1, column=c)
            if cell.value and str(cell.value).strip().lower() == old_column_name.strip().lower():
                cell.value = new_column_name
                renamed = True
                break

        if not renamed:
            raise ValueError(f"Column '{old_column_name}' not found in '{ws.title}'.")

        return f"Renamed column '{old_column_name}' to '{new_column_name}' in sheet '{ws.title}'."

    @classmethod
    def execute_replace_values(
        cls,
        wb: openpyxl.Workbook,
        sheet_name: str,
        old_val: Any,
        new_val: Any,
    ) -> str:
        """Find and replace all instances of matching value in sheet."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        replaced_count = 0

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None and str(cell.value).strip() == str(old_val).strip():
                    cell.value = new_val
                    replaced_count += 1

        return f"Replaced {replaced_count} instance(s) of '{old_val}' with '{new_val}' in '{ws.title}'."
