"""Formula operations: insert, replace, fill semantic formulas across worksheet columns."""

from typing import Optional, Dict
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from copilots_app.services.excel.utils.formula_translator import FormulaTranslator


class FormulaOps:
    """Executes formula insertions and translations on openpyxl Workbook."""

    @classmethod
    def execute_insert_formula(
        cls,
        wb: openpyxl.Workbook,
        sheet_name: str,
        column_name: str,
        semantic_formula: str,
        table_name: Optional[str] = None,
    ) -> str:
        """Insert semantic formula down a target column."""
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        min_r, max_r, min_c, max_c = 1, ws.max_row or 1, 1, ws.max_column or 1
        is_native_table = False

        if table_name and hasattr(ws, "tables") and table_name in ws.tables:
            tbl = ws.tables[table_name]
            tbl_ref = tbl if isinstance(tbl, str) else getattr(tbl, "ref", str(tbl))
            min_c, min_r, max_c, max_r = range_boundaries(tbl_ref)
            is_native_table = True

        # Build column map: lowercased header -> column letter
        col_map: Dict[str, str] = {}
        target_col_idx = None

        for c in range(min_c, max_c + 1):
            val = ws.cell(row=min_r, column=c).value
            if val is not None:
                header_name = str(val).strip()
                col_letter = get_column_letter(c)
                col_map[header_name.lower()] = col_letter
                if header_name.lower() == column_name.strip().lower():
                    target_col_idx = c

        # If target column header doesn't exist, create it at max_c + 1
        if not target_col_idx:
            target_col_idx = max_c + 1
            ws.cell(row=min_r, column=target_col_idx, value=column_name)
            col_map[column_name.strip().lower()] = get_column_letter(target_col_idx)
            if table_name and hasattr(ws, "tables") and table_name in ws.tables:
                tbl = ws.tables[table_name]
                new_ref = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(target_col_idx)}{max_r}"
                if isinstance(tbl, str):
                    ws.tables[table_name] = new_ref
                elif hasattr(tbl, "ref"):
                    tbl.ref = new_ref

        # Write formulas down data rows
        start_row = min_r + 1
        end_row = max(start_row, max_r)
        inserted_count = 0

        for r in range(start_row, end_row + 1):
            excel_formula = FormulaTranslator.translate(
                semantic_formula=semantic_formula,
                col_name_to_letter=col_map,
                row_index=r,
                is_native_table=is_native_table,
            )
            ws.cell(row=r, column=target_col_idx, value=excel_formula)
            inserted_count += 1

        return f"Inserted formula '{semantic_formula}' into column '{column_name}' across {inserted_count} row(s)."
