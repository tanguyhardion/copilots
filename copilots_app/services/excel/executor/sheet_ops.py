"""Sheet operations: create, rename, duplicate, delete worksheets."""

from typing import Dict, Any
import openpyxl


class SheetOps:
    """Executes worksheet level operations on openpyxl Workbook."""

    @classmethod
    def execute_create_sheet(cls, wb: openpyxl.Workbook, sheet_name: str) -> str:
        """Create a new worksheet."""
        if sheet_name in wb.sheetnames:
            # Generate unique name
            counter = 1
            new_name = f"{sheet_name}_{counter}"
            while new_name in wb.sheetnames:
                counter += 1
                new_name = f"{sheet_name}_{counter}"
            sheet_name = new_name

        wb.create_sheet(title=sheet_name)
        return f"Created worksheet '{sheet_name}'."

    @classmethod
    def execute_rename_sheet(cls, wb: openpyxl.Workbook, old_name: str, new_name: str) -> str:
        """Rename an existing worksheet."""
        if old_name not in wb.sheetnames:
            raise ValueError(f"Worksheet '{old_name}' not found in workbook.")

        ws = wb[old_name]
        ws.title = new_name
        return f"Renamed worksheet '{old_name}' to '{new_name}'."

    @classmethod
    def execute_duplicate_sheet(cls, wb: openpyxl.Workbook, sheet_name: str, new_name: str = None) -> str:
        """Duplicate an existing worksheet."""
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' not found.")

        source_ws = wb[sheet_name]
        new_ws = wb.copy_worksheet(source_ws)
        if new_name:
            new_ws.title = new_name

        return f"Duplicated worksheet '{sheet_name}' as '{new_ws.title}'."

    @classmethod
    def execute_delete_sheet(cls, wb: openpyxl.Workbook, sheet_name: str) -> str:
        """Delete a worksheet from workbook."""
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' not found.")

        if len(wb.sheetnames) <= 1:
            raise ValueError("Cannot delete the only remaining sheet in the workbook.")

        del wb[sheet_name]
        return f"Deleted worksheet '{sheet_name}'."
