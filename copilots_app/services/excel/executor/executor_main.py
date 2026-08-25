"""Action Executor: Central dispatcher for executing action protocols on Excel workbooks."""

import os
import openpyxl
from typing import Tuple, List, Dict, Any, Optional
from copilots_app.services.excel.models.semantic import WorkbookModel
from copilots_app.services.excel.models.protocol import (
    ActionProtocol,
    ActionIntent,
    ExecutionResult,
    ExecutionStatus,
)
from copilots_app.services.excel.validator.action_validator import ActionValidator
from copilots_app.services.excel.utils.backup_manager import BackupManager
from copilots_app.services.excel.analyzer.workbook_analyzer import WorkbookAnalyzer

from .sheet_ops import SheetOps
from .data_ops import DataOps
from .formula_ops import FormulaOps
from .style_ops import StyleOps
from .chart_ops import ChartOps
from .search_ops import SearchOps


class ActionExecutor:
    """Executes validated ActionProtocol instances against an Excel file."""

    def __init__(self, backup_dir: str = ".backups"):
        self.backup_manager = BackupManager(backup_dir=backup_dir)

    def execute(
        self,
        protocol: ActionProtocol,
        file_path: str,
        model: Optional[WorkbookModel] = None,
        save_file: bool = True,
    ) -> Tuple[ExecutionResult, Optional[WorkbookModel]]:
        """Execute action protocol on Excel workbook.

        Returns:
            Tuple of (ExecutionResult, updated_WorkbookModel)
        """
        # 1. Validate actions first
        val_res = ActionValidator.validate(protocol, model)
        if not val_res.is_valid:
            return (
                ExecutionResult(
                    status=ExecutionStatus.VALIDATION_FAILED,
                    actions_executed=0,
                    objects_modified=[],
                    warnings=val_res.warnings,
                    errors=val_res.errors,
                    details=[],
                ),
                model,
            )

        if protocol.intent == ActionIntent.NO_ACTION or not protocol.actions:
            return (
                ExecutionResult(
                    status=ExecutionStatus.SUCCESS,
                    actions_executed=0,
                    objects_modified=[],
                    warnings=["No actions specified in protocol."],
                    errors=[],
                    details=[],
                ),
                model,
            )

        # Handle read-only Query Intent
        if protocol.intent == ActionIntent.QUERY_WORKBOOK:
            return self._execute_queries(protocol, file_path, model)

        # 2. Create Safety Backup before modification
        backup_path = self.backup_manager.create_backup(file_path)

        # 3. Load openpyxl Workbook
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
        except Exception as err:
            return (
                ExecutionResult(
                    status=ExecutionStatus.EXECUTION_FAILED,
                    actions_executed=0,
                    objects_modified=[],
                    warnings=[],
                    errors=[f"Failed to open workbook for editing: {str(err)}"],
                    details=[],
                ),
                model,
            )

        actions_executed = 0
        objects_modified = set()
        warnings = list(val_res.warnings)
        errors = []
        details = []

        # 4. Dispatch each action
        for idx, item in enumerate(protocol.actions, start=1):
            act_name = item.action.lower().strip()
            sheet = item.sheet or (model.active_sheet if model else wb.active.title)

            try:
                msg = ""
                # Sheet operations
                if act_name == "create_sheet":
                    target_s = item.new_name or sheet
                    msg = SheetOps.execute_create_sheet(wb, target_s)
                    objects_modified.add(f"Sheet '{target_s}'")

                elif act_name == "rename_sheet":
                    msg = SheetOps.execute_rename_sheet(wb, sheet, item.new_name)
                    objects_modified.add(f"Sheet '{item.new_name}'")

                elif act_name == "duplicate_sheet":
                    msg = SheetOps.execute_duplicate_sheet(wb, sheet, item.new_name)
                    objects_modified.add(f"Sheet '{sheet}' duplicate")

                elif act_name == "delete_sheet":
                    msg = SheetOps.execute_delete_sheet(wb, sheet)
                    objects_modified.add(f"Sheet '{sheet}' deleted")

                # Data operations
                elif act_name in ("add_column", "add_formula_column"):
                    col_n = item.column or item.new_name or "New Column"
                    msg = DataOps.execute_add_column(
                        wb, sheet_name=sheet, column_name=col_n, table_name=item.table, default_values=item.values
                    )
                    objects_modified.add(f"Column '{col_n}' in {item.table or sheet}")

                elif act_name == "remove_column":
                    msg = DataOps.execute_remove_column(
                        wb, sheet_name=sheet, column_name=item.column, table_name=item.table
                    )
                    objects_modified.add(f"Column '{item.column}' removed")

                elif act_name == "rename_column":
                    msg = DataOps.execute_rename_column(
                        wb, sheet_name=sheet, old_column_name=item.column, new_column_name=item.new_name
                    )
                    objects_modified.add(f"Column '{item.new_name}'")

                elif act_name == "replace_values":
                    old_v = item.params.get("old_value", item.search_query)
                    new_v = item.params.get("new_value", item.new_name)
                    msg = DataOps.execute_replace_values(wb, sheet_name=sheet, old_val=old_v, new_val=new_v)
                    objects_modified.add(f"Sheet '{sheet}' values")

                # Formula operations
                elif act_name in ("insert_formula", "replace_formula", "fill_formula"):
                    msg = FormulaOps.execute_insert_formula(
                        wb, sheet_name=sheet, column_name=item.column or "Formula", semantic_formula=item.formula, table_name=item.table
                    )
                    objects_modified.add(f"Formula in {item.column or item.table or sheet}")

                # Formatting operations
                elif act_name == "autofit_columns":
                    msg = StyleOps.execute_autofit_columns(wb, sheet_name=sheet)
                    objects_modified.add(f"Sheet '{sheet}' column layout")

                elif act_name == "apply_style":
                    msg = StyleOps.execute_apply_style(
                        wb,
                        sheet_name=sheet,
                        target=item.params.get("target", "headers"),
                        bg_color=item.params.get("bg_color", "1F4E78"),
                        text_color=item.params.get("text_color", "FFFFFF"),
                        number_format=item.params.get("number_format"),
                    )
                    objects_modified.add(f"Styles in '{sheet}'")

                elif act_name == "freeze_panes":
                    cell_ref = item.params.get("cell", "A2")
                    msg = StyleOps.execute_freeze_panes(wb, sheet_name=sheet, cell_ref=cell_ref)
                    objects_modified.add(f"Freeze panes in '{sheet}'")

                # Chart operations
                elif act_name in ("create_chart", "update_chart"):
                    c_title = item.chart or "Chart"
                    c_type = item.chart_type or "bar"
                    msg = ChartOps.execute_create_chart(
                        wb, sheet_name=sheet, chart_title=c_title, chart_type=c_type, data_range=item.data_range
                    )
                    objects_modified.add(f"Chart '{c_title}'")

                elif act_name == "delete_chart":
                    msg = ChartOps.execute_delete_chart(wb, sheet_name=sheet, chart_name=item.chart)
                    objects_modified.add(f"Chart '{item.chart}' deleted")

                else:
                    msg = f"Executed custom action {act_name}."

                actions_executed += 1
                details.append({"action_id": idx, "action": act_name, "status": "success", "message": msg})

            except Exception as action_err:
                err_msg = f"Action #{idx} ({act_name}) failed: {str(action_err)}"
                errors.append(err_msg)
                details.append({"action_id": idx, "action": act_name, "status": "failed", "message": err_msg})

        # 5. Save workbook
        if save_file and actions_executed > 0:
            try:
                wb.save(file_path)
            except Exception as save_err:
                errors.append(f"Failed to save workbook file: {str(save_err)}")

        wb.close()

        # 6. Re-analyze updated workbook
        updated_model = model
        if actions_executed > 0 and save_file:
            try:
                updated_model = WorkbookAnalyzer.analyze(file_path)
            except Exception:
                pass

        final_status = (
            ExecutionStatus.SUCCESS
            if not errors
            else (ExecutionStatus.PARTIAL_SUCCESS if actions_executed > 0 else ExecutionStatus.EXECUTION_FAILED)
        )

        return (
            ExecutionResult(
                status=final_status,
                actions_executed=actions_executed,
                objects_modified=sorted(list(objects_modified)),
                warnings=warnings,
                errors=errors,
                details=details,
            ),
            updated_model,
        )

    def _execute_queries(
        self, protocol: ActionProtocol, file_path: str, model: Optional[WorkbookModel]
    ) -> Tuple[ExecutionResult, Optional[WorkbookModel]]:
        """Execute read-only search queries."""
        query_results = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for item in protocol.actions:
                act = item.action.lower().strip()
                q = item.search_query or item.column or item.sheet or ""

                if act == "find_sheet" and model:
                    query_results.extend(SearchOps.find_sheet(model, q))
                elif act == "find_table" and model:
                    query_results.extend(SearchOps.find_table(model, q))
                elif act == "find_column" and model:
                    query_results.extend(SearchOps.find_column(model, q))
                else:
                    query_results.extend(SearchOps.search_text(wb, q))
            wb.close()

            return (
                ExecutionResult(
                    status=ExecutionStatus.SUCCESS,
                    actions_executed=len(protocol.actions),
                    objects_modified=[],
                    warnings=[],
                    errors=[],
                    query_results=query_results,
                ),
                model,
            )
        except Exception as e:
            return (
                ExecutionResult(
                    status=ExecutionStatus.EXECUTION_FAILED,
                    actions_executed=0,
                    errors=[f"Query execution failed: {str(e)}"],
                ),
                model,
            )
