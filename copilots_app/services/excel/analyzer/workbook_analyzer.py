"""Workbook Analyzer: Analyzes openpyxl workbooks to extract semantic models and generate LLM prompts."""

import os
import json
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.worksheet.table import Table as OpenpyxlTable
from openpyxl.utils import range_boundaries

from copilots_app.services.excel.models.semantic import (
    WorkbookModel,
    WorksheetModel,
    TableModel,
    ColumnModel,
    NamedRangeModel,
    ChartModel,
)


class WorkbookAnalyzer:
    """Extracts semantic structure from an Excel workbook file."""

    @classmethod
    def analyze(cls, file_path: str) -> WorkbookModel:
        """Parse an Excel file and return its complete WorkbookModel."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=False)
        filename = os.path.basename(file_path)

        worksheets: List[WorksheetModel] = []
        for index, ws_name in enumerate(wb.sheetnames):
            ws = wb[ws_name]
            sheet_model = cls._analyze_worksheet(ws, index)
            worksheets.append(sheet_model)

        named_ranges: List[NamedRangeModel] = []
        try:
            for dn in wb.defined_names.definedName:
                named_ranges.append(
                    NamedRangeModel(
                        name=dn.name,
                        worksheet=dn.attr_text if hasattr(dn, "attr_text") else None,
                        value=str(dn.value),
                    )
                )
        except Exception:
            pass

        active_sheet_name = wb.active.title if wb.active else (wb.sheetnames[0] if wb.sheetnames else "")

        wb.close()

        return WorkbookModel(
            filename=filename,
            file_path=os.path.abspath(file_path),
            worksheets=worksheets,
            named_ranges=named_ranges,
            active_sheet=active_sheet_name,
        )

    @classmethod
    def _analyze_worksheet(cls, ws: openpyxl.worksheet.worksheet.Worksheet, index: int) -> WorksheetModel:
        """Analyze a single worksheet tab."""
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Count formulas
        formulas_count = 0
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas_count += 1

        tables: List[TableModel] = []
        processed_table_cells = set()

        # 1. Native Excel Tables
        if hasattr(ws, "tables"):
            for tbl_name, tbl in ws.tables.items():
                table_model = cls._analyze_native_table(ws, tbl_name, tbl)
                tables.append(table_model)
                # Track cells in this table
                min_col, min_row, max_c, max_r = range_boundaries(table_model.range)
                for r in range(min_row, max_r + 1):
                    for c in range(min_col, max_c + 1):
                        processed_table_cells.add((r, c))

        # 2. Inferred Tabular Regions (if no native tables or additional non-empty regions exist)
        if not tables and max_row > 0 and max_col > 0:
            inferred_table = cls._infer_table_region(ws, processed_table_cells)
            if inferred_table:
                tables.append(inferred_table)

        # 3. Extract Charts
        charts: List[ChartModel] = []
        if hasattr(ws, "_charts"):
            for i, c in enumerate(ws._charts):
                title = str(c.title) if c.title else f"Chart {i+1}"
                charts.append(
                    ChartModel(
                        name=title,
                        chart_type=c.__class__.__name__.replace("Chart", "").lower(),
                        title=title,
                        worksheet=ws.title,
                        cell_anchor=str(c.anchor) if hasattr(c, "anchor") else "E2",
                    )
                )

        return WorksheetModel(
            name=ws.title,
            index=index,
            is_hidden=(ws.sheet_state != "visible"),
            max_row=max_row,
            max_column=max_col,
            tables=tables,
            charts=charts,
            formulas_count=formulas_count,
        )

    @classmethod
    def _analyze_native_table(cls, ws: Any, name: str, tbl: Any) -> TableModel:
        """Extract columns and info from native openpyxl Table."""
        if isinstance(tbl, str):
            ref = tbl
        elif hasattr(tbl, "ref"):
            ref = tbl.ref
        else:
            ref = str(tbl)

        min_col, min_row, max_col, max_row = range_boundaries(ref)

        columns: List[ColumnModel] = []

        headers = []
        for c in range(min_col, max_col + 1):
            cell_val = ws.cell(row=min_row, column=c).value
            headers.append(str(cell_val) if cell_val is not None else f"Column{c - min_col + 1}")

        data_row_count = max(0, max_row - min_row)

        for col_idx, col_name in enumerate(headers, start=min_col):
            sample_vals = []
            has_formula = False
            data_type = "string"

            for r in range(min_row + 1, min(min_row + 6, max_row + 1)):
                val = ws.cell(row=r, column=col_idx).value
                if val is not None:
                    if isinstance(val, str) and val.startswith("="):
                        has_formula = True
                    sample_vals.append(val)

            if sample_vals:
                first_val = sample_vals[0]
                if isinstance(first_val, (int, float)):
                    data_type = "numeric"
                elif isinstance(first_val, str) and first_val.startswith("="):
                    data_type = "formula"

            columns.append(
                ColumnModel(
                    name=col_name,
                    index=col_idx,
                    data_type=data_type,
                    has_formula=has_formula,
                    sample_values=sample_vals,
                )
            )

        return TableModel(
            name=name,
            worksheet=ws.title,
            range=ref,
            is_native_table=True,
            columns=columns,
            row_count=data_row_count,
        )

    @classmethod
    def _infer_table_region(cls, ws: Any, processed_cells: set) -> Optional[TableModel]:
        """Infer tabular region from worksheet content when explicit tables are absent."""
        min_row, max_row = ws.min_row or 1, ws.max_row or 1
        min_col, max_col = ws.min_column or 1, ws.max_column or 1

        if max_row < 1 or max_col < 1:
            return None

        # Find header row (first non-empty row)
        header_row = min_row
        while header_row <= max_row:
            row_vals = [ws.cell(row=header_row, column=c).value for c in range(min_col, max_col + 1)]
            if any(v is not None for v in row_vals):
                break
            header_row += 1

        if header_row > max_row:
            return None

        headers = []
        columns = []
        for c in range(min_col, max_col + 1):
            val = ws.cell(row=header_row, column=c).value
            col_name = str(val).strip() if val is not None else f"Column_{c}"
            headers.append(col_name)

            sample_vals = []
            has_formula = False
            for r in range(header_row + 1, min(header_row + 6, max_row + 1)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    if isinstance(v, str) and v.startswith("="):
                        has_formula = True
                    sample_vals.append(v)

            columns.append(
                ColumnModel(
                    name=col_name,
                    index=c,
                    data_type="numeric" if sample_vals and isinstance(sample_vals[0], (int, float)) else "string",
                    has_formula=has_formula,
                    sample_values=sample_vals,
                )
            )

        range_str = f"{openpyxl.utils.get_column_letter(min_col)}{header_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

        return TableModel(
            name=f"{ws.title}Table",
            worksheet=ws.title,
            range=range_str,
            is_native_table=False,
            columns=columns,
            row_count=max(0, max_row - header_row),
        )

    @classmethod
    def generate_system_prompt(cls, model: WorkbookModel) -> str:
        """Generate formatted prompt context for the internal LLM."""
        compact = json.dumps(model.to_compact_dict(), indent=2)

        prompt = f"""SYSTEM INSTRUCTIONS: EXCEL AI COPILOT

You are an Excel AI Copilot. You analyze spreadsheet structures and propose precise workbook modifications.

CRITICAL RULES:
1. You NEVER edit Excel directly.
2. Refer to workbook elements SEMANTICALLY by Table name, Column name, or Worksheet name (NOT raw cell coordinates like B12 unless necessary).
3. Use semantic formula abstractions like `{{Profit}}/{{Revenue}}` or `SUM({{Revenue}})` instead of explicit cell coordinates.
4. When workbook interaction is required, include exactly ONE markdown block tagged with `excel-action`.

WORKBOOK CONTEXT:
```json
{compact}
```

FORMAT EXAMPLE:
I will add a Margin % column to the Sales table and update the dashboard chart.

```excel-action
{{
  "intent": "modify_workbook",
  "actions": [
    {{
      "action": "add_column",
      "table": "Sales",
      "column": "Margin %"
    }},
    {{
      "action": "insert_formula",
      "table": "Sales",
      "column": "Margin %",
      "formula": "{{Profit}}/{{Revenue}}"
    }}
  ]
}}
```
"""
        return prompt
